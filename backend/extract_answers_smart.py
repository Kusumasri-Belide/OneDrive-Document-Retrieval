import os
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))

from backend.config import DOCS_DIR, PROCESSED_DIR

# Import extraction libraries
import fitz  # PyMuPDF

try:
    from unstructured.partition.docx import partition_docx
    from unstructured.partition.ppt import partition_ppt
    from unstructured.partition.pptx import partition_pptx
    UNSTRUCTURED_AVAILABLE = True
except ImportError:
    UNSTRUCTURED_AVAILABLE = False

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def _write_txt(base_name: str, text: str):
    """Write extracted text to file."""
    out_path = os.path.join(PROCESSED_DIR, base_name + ".txt")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)

def _extract_text_generic(path: str) -> str:
    """Extract text from generic text files."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()

def _extract_xlsx(file_path: str) -> str:
    """Extract text from Excel files."""
    if not OPENPYXL_AVAILABLE:
        return ""
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"
            
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        
        wb.close()
        return text
    except Exception as e:
        print(f"Excel extraction error: {e}")
        return ""

def _join(elems):
    """Join unstructured elements."""
    return "\n".join([e.text for e in elems if getattr(e, "text", None)])

def _test_pdf_extraction(file_path):
    """Test if PDF text can actually be extracted (not just opened)."""
    try:
        # Test with PyMuPDF first
        doc = fitz.open(file_path)
        if len(doc) == 0:
            doc.close()
            return False, "Empty PDF"
        
        # Try to extract text from first page
        page = doc[0]
        test_text = page.get_text()
        doc.close()
        
        if test_text and test_text.strip():
            return True, "PyMuPDF can extract text"
        
        # If PyMuPDF fails, try pdfplumber
        if PDFPLUMBER_AVAILABLE:
            with pdfplumber.open(file_path) as pdf:
                if len(pdf.pages) > 0:
                    test_text = pdf.pages[0].extract_text()
                    if test_text and test_text.strip():
                        return True, "pdfplumber can extract text"
        
        return False, "No extractable text found"
        
    except Exception as e:
        return False, f"Extraction test failed: {str(e)}"

def _extract_pdf_safe(file_path):
    """Extract PDF text using the method that works."""
    try:
        # Try PyMuPDF first
        doc = fitz.open(file_path)
        text = ""
        for page_num in range(len(doc)):
            try:
                page = doc[page_num]
                page_text = page.get_text()
                if page_text and page_text.strip():
                    text += f"\n--- Page {page_num + 1} ---\n{page_text}"
            except:
                continue
        doc.close()
        
        if text and text.strip():
            return text
        
        # Try pdfplumber as fallback
        if PDFPLUMBER_AVAILABLE:
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text and page_text.strip():
                            text += f"\n--- Page {page_num + 1} ---\n{page_text}"
                    except:
                        continue
            return text
        
        return ""
    except Exception:
        return ""

def extract_smart():
    """Smart extraction that pre-tests files before processing."""
    print("Smart Document Extraction")
    print("=" * 50)
    
    processed_count = 0
    skipped_count = 0
    failed_count = 0
    
    # Get list of all files
    all_files = []
    for root, _, files in os.walk(DOCS_DIR):
        for fn in files:
            fp = os.path.join(root, fn)
            rel = os.path.relpath(fp, DOCS_DIR)
            all_files.append((fp, rel, fn))
    
    print(f"Found {len(all_files)} files to process\n")
    
    for fp, rel, fn in all_files:
        base = os.path.splitext(rel)[0].replace("\\", "__").replace("/", "__")
        
        # Skip if already processed
        output_path = os.path.join(PROCESSED_DIR, base + ".txt")
        if os.path.exists(output_path):
            print(f"Skipping (already processed): {rel}")
            skipped_count += 1
            continue
        
        print(f"Analyzing: {rel}")
        
        try:
            text = ""
            should_process = True
            
            if fn.lower().endswith(".pdf"):
                # Pre-test PDF extraction
                can_extract, reason = _test_pdf_extraction(fp)
                if can_extract:
                    print(f"PDF test passed: {reason}")
                    text = _extract_pdf_safe(fp)
                else:
                    print(f"PDF test failed: {reason}")
                    should_process = False
                    
            elif fn.lower().endswith(".docx") and UNSTRUCTURED_AVAILABLE:
                try:
                    # Test with a quick read
                    elements = partition_docx(filename=fp)
                    text = _join(elements)
                    print(f"DOCX extraction successful")
                except Exception as e:
                    print(f"DOCX extraction failed: {e}")
                    # Try alternative .docx.docx file
                    alt_path = fp + ".docx"
                    if os.path.exists(alt_path):
                        try:
                            print(f"Trying alternative: {alt_path}")
                            elements = partition_docx(filename=alt_path)
                            text = _join(elements)
                            print(f"Alternative DOCX successful")
                        except Exception as e2:
                            print(f"Alternative also failed: {e2}")
                            should_process = False
                    else:
                        should_process = False
                        
            elif fn.lower().endswith(".pptx") and UNSTRUCTURED_AVAILABLE:
                try:
                    elements = partition_pptx(filename=fp)
                    text = _join(elements)
                    print(f"PPTX extraction successful")
                except Exception as e:
                    print(f"PPTX extraction failed: {e}")
                    should_process = False
                    
            elif fn.lower().endswith(".ppt") and UNSTRUCTURED_AVAILABLE:
                try:
                    elements = partition_ppt(filename=fp)
                    text = _join(elements)
                    print(f"PPT extraction successful")
                except Exception as e:
                    print(f"PPT extraction failed: {e}")
                    should_process = False
                    
            elif fn.lower().endswith((".txt", ".csv", ".html")):
                text = _extract_text_generic(fp)
                print(f"Text file read successful")
                
            elif fn.lower().endswith((".xlsx", ".xls")) and OPENPYXL_AVAILABLE:
                try:
                    text = _extract_xlsx(fp)
                    if text and text.strip():
                        print(f"Excel extraction successful")
                    else:
                        print(f"Excel file appears empty")
                        should_process = False
                except Exception as e:
                    print(f"Excel extraction failed: {e}")
                    should_process = False
                
            else:
                print(f"Skipping unknown file type")
                skipped_count += 1
                continue
            
            if should_process and text and text.strip():
                _write_txt(base, text)
                print(f"Saved → {base}.txt")
                processed_count += 1
            elif should_process:
                print(f"No text content found")
                failed_count += 1
            else:
                print(f"Skipped due to extraction issues")
                failed_count += 1
                
        except Exception as e:
            print(f"Unexpected error: {e}")
            failed_count += 1
        
        print()  # Empty line for readability
    
    print("=" * 50)
    print(f"Smart Extraction Summary:")
    print(f"Successfully processed: {processed_count} files")
    print(f"Already processed/skipped: {skipped_count} files")
    print(f"Failed to process: {failed_count} files")
    print(f"Output directory: {PROCESSED_DIR}")

if __name__ == "__main__":
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    extract_smart()